from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponse
from django.db.models import Q

from .forms import SignUpForm, SignUpUpdateForm, LoginForm, ComponentForm, ActivitieForm, MasterPlanForm, DetailForm, FilterForm
from .models import CustomUser, Component, Activitie, MasterPlan, Detail, master_plan_status, detail_status

from .functions import is_admin 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Global variables
filter_records = None 

# Authentication
def SignUpView(request):
    
    if request.method == 'POST':
        form = SignUpForm(request.POST)

        password = request.POST['password']
        password_confirmation = request.POST['password_confirmation']

        if form.is_valid():
            if password != password_confirmation:
                form.add_error('password', 'Las contraseñas no coinciden.')
                
                return render(request=request, template_name='create.html', context={'form': form})

            user = CustomUser.objects.create()
            user.first_name = request.POST['first_name']
            user.last_name = request.POST['last_name']
            user.email = request.POST['email']
            user.password = make_password(password)
            user.save()

            return redirect('master-plan')
    else:
        form = SignUpForm()

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

def LoginView(request):
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        
        if form.is_valid():
            
            email = request.POST['email']
            password = request.POST['password']
            user = authenticate(request, username=email, password=password)
            
            if user is not None:
                login(request, user)
                return redirect('master-plan')
            else:
                form.add_error('password', 'Credenciales inválidas')
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})

def LogoutView(request):
    logout(request)

    return redirect('login')

# Main views
@login_required
def MasterDetailView(request, pk):

    global filter_records

    request.session['previous_url'] = request.get_full_path()

    if request.method == 'GET':
        form = FilterForm()
        
        master = MasterPlan.objects.get(id=pk)
        master_names = MasterPlan._meta.fields

        # Normal
        excluded_fields = [
            'master_plan', 
            'expected_results', 
            'objectives', 
            'goal', 
            'tasks', 
            'quantities', 
            'unit_cost',
            'total',
            'evaluation',
            'observations']

        records_name = [field for field in Detail._meta.fields if field.name not in excluded_fields]
        
        if is_admin(request.user) == True:
            records = Detail.objects.filter(master_plan=master.id)
        else:
            records = Detail.objects.filter(
                Q(goal_manager=request.user.id) |
                Q(activity_manager=request.user.id) |
                Q(supervision_manager=request.user.id),
                master_plan=master.id)

        for instance in records:
            instance.status = dict(detail_status)[instance.status]
    
    if request.method == 'POST':
        form = FilterForm(request.POST)

        master = MasterPlan.objects.get(id=pk)
        master_names = MasterPlan._meta.fields

        # Normal
        excluded_fields = [
            'master_plan', 
            'expected_results', 
            'objectives', 
            'goal', 
            'tasks', 
            'quantities', 
            'unit_cost',
            'total',
            'evaluation',
            'observations']

        records_name = [field for field in Detail._meta.fields if field.name not in excluded_fields]
        
        if is_admin(request.user) == True:
            records = Detail.objects.filter(master_plan=master.id)
        else:
            records = Detail.objects.filter(
                Q(goal_manager=request.user.id) |
                Q(activity_manager=request.user.id) |
                Q(supervision_manager=request.user.id),
                master_plan=master.id)
                    
        if form.is_valid():
            action = request.POST.get('action')
            
            if action == 'search':
                # Filters
                component = form.cleaned_data['component']
                activity = form.cleaned_data['activity']
                responsible = form.cleaned_data['responsible']
                status = form.cleaned_data['status']
                scheduled_date = form.cleaned_data['scheduled_date']
                completed_date = form.cleaned_data['completed_date']

                form = FilterForm(initial={'component':component, 'activity':activity, 'responsible':responsible, 'status':status, 'scheduled_date':scheduled_date, 'completed_date':completed_date})
                
                if component:
                    records = records.filter(component=component)
                if activity:
                    records = records.filter(activity=activity)
                if responsible:
                    records = records.filter(
                        Q(goal_manager=responsible) |
                        Q(activity_manager=responsible) |
                        Q(supervision_manager=responsible))
                if status != 'C':
                    records = records.filter(status=status)
                if scheduled_date:
                    records = records.filter(scheduled_date=scheduled_date)
                if completed_date:
                    records = records.filter(completed_date=completed_date)

                filter_records = records
            else:
                filter_records = None

            for instance in records:
                instance.status = dict(detail_status)[instance.status]

            return render(request=request, template_name='master-detail.html', context={'master':master, 'master-names':master_names, 'records':records, 'records_name':records_name, 'form':form}) 
    return render(request=request, template_name='master-detail.html', context={'master':master, 'master-names':master_names, 'records':records, 'records_name':records_name, 'form':form})

def PrintToExcelView(request, pk):
    
    records = Detail.objects.filter(master_plan=pk)

    for instance in records:
        instance.status = dict(detail_status)[instance.status]

    workbook = Workbook()
    sheet = workbook.active

    row_number = 2
    header_style = Font(name='Times New Roman', size=13, bold=True)
    row_style = Font(name='Times New Roman', size=12)
    background_color = PatternFill(start_color="52b788", end_color="52b788", fill_type="solid")

    header_row = [
        'ID', 
        'Componente',
        'Actividad',
        'Gerente de objetivo',
        'Responsable actividad',
        'Responsable supervisión',
        'Resultados esperados',
        'Objetivos',
        'Meta',
        'Tareas',
        'Estado',
        'Fecha programada', 'Fecha completada', 'Cantidades', 'Costo de unidad', 'Monto total', 'Evaluación', 'Observaciones'
    ]
    
    for column, header_cell in enumerate(header_row, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.font = header_style
        cell.value = header_cell
    
    for record in records:

        value_row = [
           str(record.id), 
           str(record.component),
           str(record.activity),
           str(record.goal_manager),
           str(record.activity_manager),
           str(record.supervision_manager),
           str(record.expected_results),
           str(record.objectives),
           str(record.goal),
           str(record.tasks),
           str(record.status),
           str(record.scheduled_date),
           str(record.completed_date),
           str(record.quantities),
           str(record.unit_cost),
           str(record.total),
           str(record.evaluation),
           str(record.observations)
        ]

        if str(record.status) == 'Completado':
            for column, value_cell in enumerate(value_row, start=1):
                cell = sheet.cell(row=row_number, column=column)
                cell.font = row_style
                cell.fill = background_color
                cell.value = value_cell
        else:
            for column, value_cell in enumerate(value_row, start=1):
                cell = sheet.cell(row=row_number, column=column)
                cell.font = row_style
                cell.value = value_cell
        
        row_number += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Plan Maestro.xlsx'
    workbook.save(response)

    return response

def PrintFilterToExcelView(request):

    global filter_records
    
    workbook = Workbook()
    sheet = workbook.active

    row_number = 2
    header_style = Font(name='Times New Roman', size=13, bold=True)
    row_style = Font(name='Times New Roman', size=12)
    background_color = PatternFill(start_color="52b788", end_color="52b788", fill_type="solid")

    header_row = [
        'ID', 
        'Componente',
        'Actividad',
        'Gerente de objetivo',
        'Responsable actividad',
        'Responsable supervisión',
        'Resultados esperados',
        'Objetivos',
        'Meta',
        'Tareas',
        'Estado',
        'Fecha programada', 'Fecha completada', 'Cantidades', 'Costo de unidad', 'Monto total', 'Evaluación', 'Observaciones'
    ]
    
    for column, header_cell in enumerate(header_row, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.font = header_style
        cell.value = header_cell
    
    if filter_records != None:
        for record in filter_records:

            value_row = [
            str(record.id), 
            str(record.component),
            str(record.activity),
            str(record.goal_manager),
            str(record.activity_manager),
            str(record.supervision_manager),
            str(record.expected_results),
            str(record.objectives),
            str(record.goal),
            str(record.tasks),
            str(record.status),
            str(record.scheduled_date),
            str(record.completed_date),
            str(record.quantities),
            str(record.unit_cost),
            str(record.total),
            str(record.evaluation),
            str(record.observations)
            ]

            if str(record.status) == 'Completado':
                for column, value_cell in enumerate(value_row, start=1):
                    cell = sheet.cell(row=row_number, column=column)
                    cell.font = row_style
                    cell.fill = background_color
                    cell.value = value_cell
            else:
                for column, value_cell in enumerate(value_row, start=1):
                    cell = sheet.cell(row=row_number, column=column)
                    cell.font = row_style
                    cell.value = value_cell
            
            row_number += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Plan Maestro.xlsx'
        workbook.save(response)
    else:
        message = "Debes de filtrar la información antes de imprimir un filtro"
        script = f"<script>alert('{message}'); window.location.href = document.referrer;</script>"

        return HttpResponse(script)

    return response

# Component Logic Views
@login_required
def ListComponentView(request):
    request.session['previous_url'] = request.get_full_path()

    records = Component.objects.all()
    records_name = Component._meta.fields

    return render(request=request, template_name='list.html', context={'records': records, 'records_name': records_name})

@login_required
def CreateComponentView(request):

    if request.method == 'POST':
        form = ComponentForm(request.POST)

        if form.is_valid():
            try:
                component = Component.objects.create(name=request.POST['name'])
                
                return redirect(request.session.get('previous_url', '/'))
            except IntegrityError:
                form.add_error('name', 'El componente ya existe.')
    else:
        form = ComponentForm()

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def UpdateComponentView(request, pk):

    component = Component.objects.get(id=pk)
    
    if request.method == 'POST':
        form = ComponentForm(request.POST)

        if form.is_valid():
            try:
                component.unique_error_message
                component.name = request.POST['name']
                component.save()
                
                return redirect(request.session.get('previous_url', '/'))
            except IntegrityError:
                form.add_error('name', 'El componente ya existe.')
    else:
        form = ComponentForm(initial={'name':component.name})

    return render(request=request, template_name='update.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def DeleteComponentView(request, pk):

    record = Component.objects.get(id=pk)
    
    if request.POST:
        record.delete()

        return redirect(request.session.get('previous_url', '/'))

    return render(request=request, template_name='delete.html', context={'previous_url':request.session.get('previous_url', '/')})

# Activity Logic Views
@login_required
def ListActivityView(request):
    request.session['previous_url'] = request.get_full_path()

    records = Activitie.objects.all()
    records_name = Activitie._meta.fields

    return render(request=request, template_name='list.html', context={'records': records, 'records_name': records_name})

@login_required
def CreateActivityView(request):

    if request.method == 'POST':
        form = ActivitieForm(request.POST)

        if form.is_valid():
            try:
                component = Component.objects.get(id=request.POST['component'])
                activity = Activitie.objects.create(component=component, name=request.POST['name'])
            
                return redirect(request.session.get('previous_url', '/'))
            except IntegrityError:
                form.add_error('name', 'La actividad ya existe.')
    else:
        form = ActivitieForm()

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def UpdateActivityView(request, pk):

    activity = Activitie.objects.get(id=pk)
    
    if request.method == 'POST':
        form = ActivitieForm(request.POST)

        if form.is_valid():
            try:
                component = Component.objects.get(id=request.POST['component'])

                activity.component = component
                activity.name = request.POST['name']
                activity.save()

                return redirect(request.session.get('previous_url', '/'))
            except IntegrityError:
                form.add_error('name', 'La actividad ya existe.')
    else:
        form = ActivitieForm(initial={'component':activity.component, 'name':activity.name})

    return render(request=request, template_name='update.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def DeleteActivityView(request, pk):

    record = Activitie.objects.get(id=pk)
    
    if request.POST:
        record.delete()

        return redirect(request.session.get('previous_url', '/'))

    return render(request=request, template_name='delete.html', context={'previous_url':request.session.get('previous_url', '/')})

# Responsible Logic Views
@login_required
def ListResponsibleView(request):
    request.session['previous_url'] = request.get_full_path()

    records = CustomUser.objects.all()
    records_name = CustomUser._meta.fields

    excluded_fields = ['is_staff', 'is_active', 'is_superuser', 'date_joined', 'last_login', 'password', 'username', 'role']

    records = CustomUser.objects.all()
    records_name = [field for field in CustomUser._meta.fields if field.name not in excluded_fields]

    return render(request=request, template_name='list.html', context={'records': records, 'records_name': records_name})

@login_required
def CreateResponsibleView(request):

    if request.method == 'POST':
        form = SignUpForm(request.POST)
    
        password = request.POST['password']
        password_confirmation = request.POST['password_confirmation']
        email = request.POST['email']

        if form.is_valid():
            if password != password_confirmation:
                form.add_error('password', 'Las contraseñas no coinciden.')
                
                return render(request=request, template_name='create.html', context={'form': form})

            if CustomUser.objects.filter(email=email).exists():
                form.add_error('email', 'El correo electronico ya esta en uso')
            else:
                responsible = CustomUser.objects.create()
                responsible.first_name = request.POST['first_name']
                responsible.last_name = request.POST['last_name']
                responsible.email = request.POST['email']
                responsible.role = 'B'
                responsible.password = make_password(password)
                responsible.save()

                return redirect(request.session.get('previous_url', '/'))
    else:
        form = SignUpForm()

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def UpdateResponsibleView(request, pk):

    responsible = CustomUser.objects.get(id=pk)
    
    if request.method == 'POST':
        form = SignUpUpdateForm(request.POST)

        password = request.POST['password']
        password_confirmation = request.POST['password_confirmation']

        if form.is_valid():
            if password_confirmation != '' and password != '' and password != password_confirmation:
                form.add_error('password', 'Las contraseñas no coinciden.')
                
                return render(request=request, template_name='create.html', context={'form': form})
            
            if password != '' and password_confirmation != '':
                responsible.password = make_password(password)

            responsible.first_name = request.POST['first_name']
            responsible.last_name = request.POST['last_name']
            responsible.email = request.POST['email']
            responsible.save()

            return redirect(request.session.get('previous_url', '/'))
    else:
        form = SignUpUpdateForm(initial={
            'first_name':responsible.first_name,
            'last_name': responsible.last_name,
            'email': responsible.email,
            })

    return render(request=request, template_name='update.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def DeleteResponsibleView(request, pk):

    record = CustomUser.objects.get(id=pk)
    
    if request.POST:
        record.delete()

        return redirect(request.session.get('previous_url', '/'))

    return render(request=request, template_name='delete.html', context={'previous_url':request.session.get('previous_url', '/')})

# MasterPlan Logic Views
@login_required
def ListMasterPlanView(request):
    request.session['previous_url'] = request.get_full_path()

    # records = MasterPlan.objects.all()
    records_name = MasterPlan._meta.fields

    if is_admin(request.user) == True:
        records = MasterPlan.objects.all()
    else:
        records = MasterPlan.objects.filter(
            Q(detail__goal_manager_id=request.user.id) |
            Q(detail__activity_manager_id=request.user.id) |
            Q(detail__supervision_manager_id=request.user.id)
        ).distinct()

    for instance in records:
        instance.status = dict(master_plan_status)[instance.status]

    return render(request=request, template_name='list.html', context={'records': records, 'records_name': records_name})

@login_required
def CreateMasterPlanView(request):

    if request.method == 'POST':
        form = MasterPlanForm(request.POST)

        if form.is_valid():
            try:
                master_plan = MasterPlan.objects.create()
                master_plan.name = request.POST['name']
                master_plan.description = request.POST['description']
                master_plan.notes = request.POST['notes']
                master_plan.save()

                return redirect(request.session.get('previous_url', '/'))
            except IntegrityError:
                form.add_error('name', 'El master plan ya existe.')
    else:
        form = MasterPlanForm()

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def UpdateMasterPlanView(request, pk):

    master_plan = MasterPlan.objects.get(id=pk)
    
    if request.method == 'POST':
        form = MasterPlanForm(request.POST)

        if form.is_valid():
            try:
                master_plan.name = request.POST['name']
                master_plan.description = request.POST['description']
                master_plan.status = request.POST['status']
                master_plan.notes = request.POST['notes']
                master_plan.save()

                return redirect(request.session.get('previous_url', '/'))
            except IntegrityError:
                form.add_error('name', 'El master plan ya existe.')
    else:
        form = MasterPlanForm(initial={
            'name':master_plan.name,
            'description':master_plan.description,
            'status':master_plan.status,
            'notes':master_plan.notes
            })

    return render(request=request, template_name='update.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def DeleteMasterPlanView(request, pk):

    record = MasterPlan.objects.get(id=pk)
    
    if request.POST:
        record.delete()

        return redirect(request.session.get('previous_url', '/'))

    return render(request=request, template_name='delete.html', context={'previous_url':request.session.get('previous_url', '/')})

# Detail Logic Views
@login_required
def ListDetailView(request):
    request.session['previous_url'] = request.get_full_path()

    records = Detail.objects.all()

    excluded_fields = [
        'master_plan', 
        'expected_results', 
        'objectives', 
        'goal', 
        'tasks', 
        'quantities', 
        'unit_cost',
        'total',
        'evaluation',
        'observations'
    ]

    records_name = [field for field in Detail._meta.fields if field.name not in excluded_fields]

    for instance in records:
        instance.status = dict(detail_status)[instance.status]

    return render(request=request, template_name='list.html', context={'records': records, 'records_name': records_name})

@login_required
def DetailInfoView(request, pk):

    if request.method == 'GET':

        # Normal
        excluded_fields = ['master_plan']

        records_name = [field for field in Detail._meta.fields if field.name not in excluded_fields]
        record = Detail.objects.get(id=pk)

        record.status = dict(detail_status)[record.status]
    
    return render(request=request, template_name='detail.html', context={'record':record, 'records_name':records_name, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def CreateDetailView(request):

    if request.method == 'POST':
        form = DetailForm(request.POST)

        if form.is_valid():
            master_plan = MasterPlan.objects.get(id=request.POST['master_plan'])
            activity = Activitie.objects.get(id=request.POST['activity'])
            component = Component.objects.get(id=activity.component.id)
            goal_manager = CustomUser.objects.get(id=request.POST['goal_manager'])
            activity_manager = CustomUser.objects.get(id=request.POST['activity_manager'])
            supervision_manager = CustomUser.objects.get(id=request.POST['supervision_manager'])
            complete_date = request.POST['completed_date']
            quantities = request.POST['quantities']
            unit_cost = request.POST['unit_cost']
            
            if complete_date == '':
                complete_date = None
            if quantities == '':
                quantities = None
            if unit_cost == '':
                unit_cost = None
            
            detail = Detail.objects.create(
                master_plan=master_plan,
                component=component,
                activity=activity,
                goal_manager=goal_manager,
                activity_manager=activity_manager,
                supervision_manager=supervision_manager,
                expected_results=request.POST['expected_results'],
                objectives=request.POST['objectives'],
                goal=request.POST['goal'],
                tasks=request.POST['tasks'],
                status=request.POST['status'],
                scheduled_date=request.POST['scheduled_date'],
                completed_date=complete_date,
                quantities=quantities,
                unit_cost=unit_cost,
                evaluation = request.POST['evaluation'],
                observations = request.POST['observations']
            )

            return redirect(request.session.get('previous_url', '/'))
    else:
        form = DetailForm()

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def CreateMasterDetailView(request, pk):

    master = MasterPlan.objects.get(id=pk)

    if request.method == 'POST':
        form = DetailForm(request.POST)

        if form.is_valid():
            master_plan = MasterPlan.objects.get(id=request.POST['master_plan'])
            activity = Activitie.objects.get(id=request.POST['activity'])
            component = Component.objects.get(id=activity.component.id)
            goal_manager = CustomUser.objects.get(id=request.POST['goal_manager'])
            activity_manager = CustomUser.objects.get(id=request.POST['activity_manager'])
            supervision_manager = CustomUser.objects.get(id=request.POST['supervision_manager'])
            complete_date = request.POST['completed_date']
            quantities = request.POST['quantities']
            unit_cost = request.POST['unit_cost']
            
            if goal_manager == '':
                complete_date = None
            if activity_manager == '':
                quantities = None
            if supervision_manager == '':
                unit_cost = None
            if complete_date == '':
                complete_date = None
            if quantities == '':
                quantities = None
            if unit_cost == '':
                unit_cost = None
            
            detail = Detail.objects.create(
                master_plan=master_plan,
                component=component,
                activity=activity,
                goal_manager=goal_manager,
                activity_manager=activity_manager,
                supervision_manager=supervision_manager,
                expected_results=request.POST['expected_results'],
                objectives=request.POST['objectives'],
                goal=request.POST['goal'],
                tasks=request.POST['tasks'],
                status=request.POST['status'],
                scheduled_date=request.POST['scheduled_date'],
                completed_date=complete_date,
                quantities=quantities,
                unit_cost=unit_cost,
                evaluation = request.POST['evaluation'],
                observations = request.POST['observations']
            )
            return redirect(request.session.get('previous_url', '/'))
    else:
        form = DetailForm(initial={
            'master_plan':master.id
        })

    return render(request=request, template_name='create.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def UpdateDetailView(request, pk):

    detail = Detail.objects.get(id=pk)
    
    if request.method == 'POST':
        form = DetailForm(request.POST)

        if form.is_valid():
            master_plan = MasterPlan.objects.get(id=request.POST['master_plan'])
            activity = Activitie.objects.get(id=request.POST['activity'])
            component = Component.objects.get(id=activity.component.id)
            goal_manager = CustomUser.objects.get(id=request.POST['goal_manager'])
            activity_manager = CustomUser.objects.get(id=request.POST['activity_manager'])
            supervision_manager = CustomUser.objects.get(id=request.POST['supervision_manager'])
            completed_date = request.POST['completed_date']
            quantities = request.POST['quantities']
            unit_cost = request.POST['unit_cost']
            
            if completed_date == '':
                completed_date = None
            if quantities == '':
                quantities = None
            if unit_cost == '':
                unit_cost = None

            detail.master_plan=master_plan
            detail.component=component
            detail.activity=activity
            detail.goal_manager=goal_manager
            detail.activity_manager=activity_manager
            detail.supervision_manager=supervision_manager
            detail.expected_results=request.POST['expected_results']
            detail.objectives=request.POST['objectives']
            detail.goal=request.POST['goal']
            detail.tasks=request.POST['tasks']
            detail.status = request.POST['status']
            detail.scheduled_date = request.POST['scheduled_date']
            detail.completed_date = completed_date
            detail.quantities=quantities
            detail.unit_cost=unit_cost
            detail.evaluation = request.POST['evaluation']
            detail.observations = request.POST['observations']
            detail.save()

            return redirect(request.session.get('previous_url', '/'))
    else:
        try:
            completed_date = detail.completed_date.strftime('%Y-%m-%d')
        except AttributeError:
            completed_date = None
        
        form = DetailForm(initial={
            'master_plan':detail.master_plan,
            'component':detail.component,
            'activity':detail.activity,
            'goal_manager':detail.goal_manager,
            'activity_manager':detail.activity_manager,
            'supervision_manager':detail.supervision_manager,
            'expected_results':detail.expected_results,
            'objectives':detail.objectives,
            'goal':detail.goal,
            'tasks':detail.tasks,
            'status':detail.status,
            'scheduled_date':detail.scheduled_date.strftime('%Y-%m-%d'),
            'completed_date':completed_date,
            'quantities':detail.quantities,
            'unit_cost':detail.unit_cost,
            'evaluation':detail.evaluation,
            'observations':detail.observations
            })

    return render(request=request, template_name='update.html', context={'form': form, 'previous_url':request.session.get('previous_url', '/')})

@login_required
def DeleteDetailView(request, pk):

    record = Detail.objects.get(id=pk)
    
    if request.POST:
        record.delete()

        return redirect(request.session.get('previous_url', '/'))

    return render(request=request, template_name='delete.html', context={'previous_url':request.session.get('previous_url', '/')})